from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
import pytest

from dbf_bridge.converters import (
    EXCEL_MAX_COLUMNS,
    EXCEL_MAX_STRING_LENGTH,
    ConversionCancelled,
    JsonlConversionError,
    jsonl_to_csv,
    jsonl_to_json,
    jsonl_to_xlsx,
)
from dbf_bridge.verifier import count_json_records


def _write_jsonl(path: Path, records: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as outfile:
        for record in records:
            outfile.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")))
            outfile.write("\n")


def test_json_streams_empty_lines_and_unicode(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.json"
    source.write_bytes(
        b'\n {"id":1,"text":"za\xc5\xbc\xc3\xb3\xc5\x82\xc4\x87 \xf0\x9f\x98\x80"}\n\n'
        b'{"id":2,"text":"a,b;\\n\\\"cytat\\\""}\n'
    )

    stats = jsonl_to_json(source, destination)

    assert stats.record_count == 2
    assert stats.engine == "binary-stream"
    assert json.loads(destination.read_text(encoding="utf-8")) == [
        {"id": 1, "text": "zażółć 😀"},
        {"id": 2, "text": 'a,b;\n"cytat"'},
    ]
    assert b",\n]" not in destination.read_bytes()
    assert count_json_records(destination) == (2, [])


def test_json_invalid_line_is_private_and_atomic(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.json"
    secret = "bardzo-tajna-wartosc"
    source.write_text(f'{{"ok":1}}\n{{"secret":"{secret}"\n', encoding="utf-8")
    destination.write_text("old", encoding="utf-8")

    with pytest.raises(JsonlConversionError) as error:
        jsonl_to_json(source, destination)

    assert "line 2" in str(error.value)
    assert secret not in str(error.value)
    assert destination.read_text(encoding="utf-8") == "old"
    assert not (tmp_path / "output.json.partial").exists()


def test_csv_polars_streaming_with_missing_and_new_fields(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.csv"
    _write_jsonl(
        source,
        [
            {"id": 1, "name": "Ala, Kowalska", "active": True, "amount": 1},
            {"id": 2, "name": 'Jan "Nowak"', "active": False, "extra": "późne"},
            {"id": 3, "name": None, "amount": 2.5},
        ],
    )

    stats = jsonl_to_csv(source, destination, separator=";")

    assert stats.engine == "polars-streaming"
    with destination.open("r", encoding="utf-8", newline="") as infile:
        rows = list(csv.DictReader(infile, delimiter=";"))
    assert list(rows[0]) == ["id", "name", "active", "amount", "extra"]
    assert rows[0]["name"] == "Ala, Kowalska"
    assert rows[1]["name"] == 'Jan "Nowak"'
    assert rows[1]["extra"] == "późne"
    assert rows[2]["name"] == ""
    assert rows[2]["amount"] == "2.5"


def test_csv_known_schema_uses_single_streaming_pass(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.csv"
    _write_jsonl(source, [{"id": 1, "name": "Ala"}, {"id": 2, "name": None}])

    stats = jsonl_to_csv(
        source,
        destination,
        columns=["id", "name"],
        schema_types={"id": "integer", "name": "string"},
        expected_record_count=2,
        source_is_validated=True,
    )

    assert stats.engine == "polars-streaming"
    assert stats.record_count == 2
    with destination.open("r", encoding="utf-8", newline="") as infile:
        assert list(csv.reader(infile)) == [["id", "name"], ["1", "Ala"], ["2", ""]]


def test_csv_nested_values_are_compact_json_or_flattened(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    compact = tmp_path / "compact.csv"
    flattened = tmp_path / "flattened.csv"
    _write_jsonl(
        source,
        [
            {
                "id": 1,
                "customer": {"name": "Jan", "city": "Warszawa"},
                "tags": ["a", "b"],
            }
        ],
    )

    compact_stats = jsonl_to_csv(source, compact)
    flattened_stats = jsonl_to_csv(source, flattened, flatten=True)

    assert compact_stats.engine == "python-stream"
    assert flattened_stats.engine == "python-stream"
    with compact.open("r", encoding="utf-8", newline="") as infile:
        compact_row = next(csv.DictReader(infile))
    assert json.loads(compact_row["customer"]) == {"name": "Jan", "city": "Warszawa"}
    assert json.loads(compact_row["tags"]) == ["a", "b"]
    with flattened.open("r", encoding="utf-8", newline="") as infile:
        flat_row = next(csv.DictReader(infile))
    assert flat_row["customer.name"] == "Jan"
    assert flat_row["customer.city"] == "Warszawa"


def test_xlsx_splits_sheets_without_losing_boundary_records(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.xlsx"
    records = [
        {
            "id": index,
            "text": "żółw 😀",
            "formula": "=SUM(A1:A10)",
            "nested": {"index": index},
        }
        for index in range(1, 8)
    ]
    _write_jsonl(source, records)

    stats = jsonl_to_xlsx(
        source,
        destination,
        columns=["id", "text", "formula", "nested"],
        max_rows_per_sheet=4,
    )

    assert stats.sheet_count == 3
    assert stats.record_count == 7
    workbook = openpyxl.load_workbook(destination, read_only=False, data_only=False)
    assert workbook.sheetnames == ["Dane_1", "Dane_2", "Dane_3"]
    ids: list[int] = []
    for worksheet in workbook.worksheets:
        assert [cell.value for cell in worksheet[1]] == ["id", "text", "formula", "nested"]
        ids.extend(
            row[0].value for row in worksheet.iter_rows(min_row=2) if row[0].value is not None
        )
    assert ids == list(range(1, 8))
    assert workbook["Dane_1"]["C2"].value == "=SUM(A1:A10)"
    assert workbook["Dane_1"]["C2"].data_type == "s"
    assert json.loads(workbook["Dane_1"]["D2"].value) == {"index": 1}
    workbook.close()


def test_xlsx_rejects_cell_and_column_limits_atomically(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.xlsx"
    _write_jsonl(source, [{"text": "x" * (EXCEL_MAX_STRING_LENGTH + 1)}])

    with pytest.raises(JsonlConversionError, match="cell limit"):
        jsonl_to_xlsx(source, destination, columns=["text"])
    assert not destination.exists()
    assert not (tmp_path / "output.xlsx.partial").exists()

    source.write_text("", encoding="utf-8")
    with pytest.raises(JsonlConversionError, match="at most"):
        jsonl_to_xlsx(
            source,
            destination,
            columns=[f"c{index}" for index in range(EXCEL_MAX_COLUMNS + 1)],
        )


def test_cancellation_removes_partial_output(tmp_path: Path) -> None:
    source = tmp_path / "input.jsonl"
    destination = tmp_path / "output.json"
    _write_jsonl(source, [{"id": index} for index in range(5000)])

    with pytest.raises(ConversionCancelled):
        jsonl_to_json(source, destination, cancel_callback=lambda: True)

    assert not destination.exists()
    assert not (tmp_path / "output.json.partial").exists()
